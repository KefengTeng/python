#!/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import *
import ipaddr
import re

# 打开Excel文件
wb = load_workbook(
    r'C:\\Users\\PaintMyLove\\Desktop\\副本IP地址占用前后台数据提取--企业专线独享1.xlsx')

# 读取源工作表
wbr = wb['SQL Results']

# 创建新工作表
wbw = wb.create_sheet('Sheet2')

# 写表头
wbw.append(['接入号', '多余的IP'])

# 数据字典
info_dict = {}

# 读取源工作表数据, 忽略表头
for row in wbr.iter_rows(min_row=2):
    # 接入号
    No = row[3].value
    # 单IP
    SingleIp = row[1].value
    # 地址段
    Net = row[5].value
    # 单IP集合
    info_dict.setdefault(No, {}).setdefault('SpSet', set()).add(SingleIp)
    # 地址段集合
    if re.search(r'^\d+.\d+.\d+.\d+\/\d+$', Net):
        info_dict.setdefault(No, {}).setdefault('NetSet', set()).add(Net)

# 遍历字典
for k, v in info_dict.items():
    # 遍历IP
    for Sp in v['SpSet']:
        # 遍历网段
        if 'NetSet' in v:
            for Nt in v['NetSet']:
                if ipaddr.IPAddress(Sp) in ipaddr.IPv4Network(Nt):
                    break
            else:
                wbw.append([k, Sp])

# 保存工作簿
wb.save(r'C:\\Users\\PaintMyLove\\Desktop\\副本IP地址占用前后台数据提取--企业专线独享1.xlsx')
