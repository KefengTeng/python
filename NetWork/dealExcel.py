#!/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import *

# 打开Excel文件
wb = load_workbook(r'C:\\Users\\PaintMyLove\\Desktop\\明细.xlsx')

# 读取源工作表
wbr = wb.get_sheet_by_name('Sheet1')

# 创建新工作表
wbw = wb.create_sheet('Sheet2')

# 数据字典
info_dict = {}

# 类型元组
type_tuple = ('SR设备IP', 'SR/MSE的Trunk口', '外层VLANID',
              '外层VLAN', 'VLANID', 'IP地址')

# 读取源工作表数据, 忽略表头
for row in wbr.iter_rows(min_row=2):
    # 接入号
    No = row[0].value
    # 字段值
    Value = row[5].value
    # 数据保存至字典
    for Type in type_tuple:
        if row[4].value == Type:
            info_dict.setdefault(No, {}).setdefault(Type, []).append(Value)
        else:
            info_dict.setdefault(No, {}).setdefault(Type, [])

# 写表头
wbw.append(['接入号', 'SR设备IP', 'SR/MSE的Trunk口',
            '外层VLANID', '外层VLAN', 'VLANID', 'IP地址'])
# 写数据
for k, v in info_dict.items():
    wbw.append([k, str(v['SR设备IP']), str(v['SR/MSE的Trunk口']),
                str(v['外层VLANID']), str(v['外层VLAN']), str(v['VLANID']), str(v['IP地址'])])

# 保存工作簿
wb.save(r'C:\\Users\\PaintMyLove\\Desktop\\明细.xlsx')
