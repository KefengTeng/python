#!/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import *

# 打开源文件
wbr = load_workbook('/home/brokensmile/Data/Python3/Label/devices.xlsx')

# 创建一个工作薄
wbw = Workbook()

# 分别读取源文件工作表并处理数据
for sheet in wbr.get_sheet_names():
    # 打开工作表
    wbt = wbr.get_sheet_by_name(sheet)
    # 在新建的工作薄内创建同名工作表
    cur_sheet = wbw.create_sheet(sheet, -1)

    irow = []
    # 读取表数据
    for row in wbt.values:
        # 过滤表头
        if row[0] != '序号':
            # 去掉单元格中内容前后存在的空格或换行符
            nrow = tuple(map(lambda s: str(s).strip(), row))
            # 处理后的数据,每5条记录保存为一行
            irow.append('工程名称1:%s\n设备型号:%s\n设备名:%s\n设备用途:%s\n设备所属节点:%s\n设备IP地址:%s\n设备序列号:%s'
                        % tuple(nrow[i] for i in range(1, len(nrow))))
            if len(irow) == 5:
                cur_sheet.append(irow)
                irow = []
# 保存工作薄
wbw.save('/home/brokensmile/Data/Python3/Label/Label.xlsx')
