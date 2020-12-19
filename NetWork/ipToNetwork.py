#!/bin/env python3
# -*- coding: utf-8 -*-

from netaddr import *
from openpyxl import *

# 打开Excel文件
wbr = load_workbook(r'C:\\Users\\PaintMyLove\\Desktop\\副本专线机房数据1218.xlsx')

# 创建一个工作薄
wbw = Workbook()

# 读取源工作表
wbs = wbr.get_sheet_by_name('机房IP地址段信息')

# 创建工作表
cur_sheet = wbw.create_sheet('机房IP地址段反掩码', -1)

# 写表头
cur_sheet.append(['*机房名称', '*起始IP地址', '*终止IP地址', '*网络号', '*反掩码'])

# 读取源工作簿数据, 忽略表头
for row in wbs.iter_rows(min_row=2):
    # 网络号
    netWork = str(iprange_to_cidrs(row[1].value, row[2].value)[0].network)
    # 反掩码
    reversedMask = str(iprange_to_cidrs(
        row[1].value, row[2].value)[0].hostmask)
    # 写工作簿
    cur_sheet.append([row[0].value, row[1].value, row[2].value, netWork, reversedMask])

# 保存工作簿
wbw.save(r'C:\\Users\\PaintMyLove\\Desktop\\转换结果.xlsx')
